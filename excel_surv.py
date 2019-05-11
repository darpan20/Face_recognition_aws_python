def excel(l):
 import openpyxl
 import os
 from datetime import date,datetime
 m=["Date","time","NAME"]
 now=date.today()
 now=now.strftime('%d-%m-%Y')
 n=datetime.now()
 t = n.strftime("%H:%M:%S")
 
 folderName = "DETECTED"
 folderPath = os.path.join(os.path.dirname(os.path.realpath('__file__')), folderName)
 wb=openpyxl.Workbook()
 i=0
 sheet = wb.active
 if not os.path.exists(folderPath):
    os.makedirs(folderPath)
    
 try:
  wb = openpyxl.load_workbook(folderPath+"\\"+"DETECTED.xlsx")
 except:
   wb.save(folderPath+"\\"+"DETECTED.xlsx") 
   while(i<3):
     x = sheet.cell(row = 1, column = i+1)
     x.value =m[i]
     i+=1
   
     print(" HEADINGS ENTERED")
 j=1
 sheet = wb.active
 s=sheet.max_row+1
 print(s)
 while(j<4):
  y = sheet.cell(row =s, column = j)
  if j==1:
    y.value=now
    j+=1
  elif j==2:
    y.value=t
    j+=1  
  else:
    y.value=l
    print(l)
    j+=1
 print("value added")
     
         
 print("end")    
 
 
 
 
 
 wb.save(folderPath+"\\"+"DETECTED.xlsx")
 return   

if __name__=='__main__':
 excel(l)
 