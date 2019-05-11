
def excel(l):
 import openpyxl
 import os
 from datetime import date
 from winsound import PlaySound,SND_ASYNC
 from Recognition_menu import app
 
 now=date.today()
 now=now.strftime('%d-%m-%Y')
 folderName = "Attendance"
 folderPath = os.path.join(os.path.dirname(os.path.realpath('__file__')), folderName)
 wb=openpyxl.Workbook()
 if not os.path.exists(folderPath):
    os.makedirs(folderPath)
 try:
  wb = openpyxl.load_workbook(folderPath+"\\"+"Attendance_Register.xlsx")
 except:
   wb.save(folderPath+"\\"+"Attendance_Register.xlsx") 
   wb = openpyxl.load_workbook(folderPath+"\\"+"Attendance_Register.xlsx")
 sheet = wb.active
 i=0
 while(True):
    x = sheet.cell(row = 1, column = i+1)
   
    j = 2
  
    if x.value == now:
        for a in range(len(l)):
            while(True):
             y=sheet.cell(row = j, column = i+1)   
             if y.value !=None and y.value!=l[a]:
              j+=1
             elif y.value==l[a]:
               j=2  
               break  
             else:
                 
               y.value = l[a]
               j=2
               break
        
        try:
           wb.save(folderPath+"\\"+"Attendance_Register.xlsx")
           print("Old Data Updated") 
        except:
            print("abc")
            PlaySound(None,SND_ASYNC)
            play = lambda: PlaySound(os.path.join(os.path.dirname(os.path.realpath('__file__')),"err.wav"),SND_ASYNC)
            play()
            app()
        return    
    if x.value!=now and x.value is not None:
      i+=1  
    else:
        x.value = now
        for a in range(len(l)):
            y = sheet.cell(row = j, column = i+1)
            y.value = l[a]
            j+=1
        try:    
         wb.save(folderPath+"\\"+"Attendance_Register.xlsx")
         print("New Data Entered!") 
        except:
            print("abc")
            PlaySound(None,SND_ASYNC)
            play = lambda: PlaySound(os.path.join(os.path.dirname(os.path.realpath('__file__')),"err.wav"),SND_ASYNC)
            play()
            app()
        
        return 

if __name__=='__main__':
    
 excel(l)
 